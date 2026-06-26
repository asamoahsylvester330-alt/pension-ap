import pandas as pd
import requests
import time
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from rapidfuzz import fuzz

from banks import get_bank_code


MATCH_THRESHOLD = 90
REVIEW_THRESHOLD = 70

COLUMN_ALIASES = {
    "submitted_name": [
        "ACCOUNT NAME",
        "submitted_name",
        "account_holder_name",
        "account holder name",
        "account_holder",
        "account holder",
        "holder_name",
        "holder name",
        "name",
        "full_name",
        "full name",
        "employee_name",
        "employee name",
        "customer_name",
        "customer name"
    ],
    "account_number": [
        "ACCOUNT NUMBER",
        "ACCOUNT NO.",
        "account_number",
        "account number",
        "account_no",
        "account no",
        "acct_no",
        "acct no",
        "acct number",
        "bank_account",
        "bank account",
        "bank_account_number",
        "bank account number"
    ],
    "bank": [
        "BANK",
        "bank",
        "bank_name",
        "bank name",
        "bankcode",
        "bank code",
        "financial institution",
        "institution"
    ]
}


def normalize_name(name):
    if pd.isna(name):
        return ""
    name = str(name).upper()
    name = re.sub(r"[^\w\s]", "", name)
    name = re.sub(r"\s+", " ", name)
    return name.strip()


def normalize_column_name(name):
    name = str(name).strip().lower()
    name = name.replace("_", " ")
    name = re.sub(r"\s+", " ", name)
    return name


def detect_column(df_columns, aliases):
    normalized_columns = {normalize_column_name(c): c for c in df_columns}
    for alias in aliases:
        alias_norm = normalize_column_name(alias)
        if alias_norm in normalized_columns:
            return normalized_columns[alias_norm]
    return None


def classify_score(score):
    if score >= MATCH_THRESHOLD:
        return "MATCH"
    elif score >= REVIEW_THRESHOLD:
        return "REVIEW"
    return "MISMATCH"


def resolve_account(account_number, bank_code, secret_key):
    headers = {"Authorization": f"Bearer {secret_key}"}
    url = (
        "https://api.paystack.co/bank/resolve"
        f"?account_number={account_number}"
        f"&bank_code={bank_code}"
    )

    try:
        response = requests.get(url, headers=headers, timeout=20)
        data = response.json()

        if response.status_code == 200 and data.get("status"):
            return {
                "success": True,
                "account_name": data["data"]["account_name"],
                "error": None
            }

        print(f"[PAYSTACK ERROR] {account_number} | {data.get('message')}")
        return {
            "success": False,
            "account_name": None,
            "error": data.get("message")
        }

    except requests.exceptions.ConnectionError:
        print(f"[NO INTERNET] {account_number}")
        return {"success": False, "account_name": None, "error": "NO_INTERNET"}

    except requests.exceptions.Timeout:
        print(f"[TIMEOUT] {account_number}")
        return {"success": False, "account_name": None, "error": "TIMEOUT"}

    except Exception as e:
        print(f"[REQUEST ERROR] {account_number} | {e}")
        return {"success": False, "account_name": None, "error": str(e)}


def read_file(input_path):
    if input_path.endswith(".csv"):
        return pd.read_csv(input_path, dtype=str)
    return pd.read_excel(input_path, dtype=str)


# =========================
# USER-FACING ERROR MESSAGES
# =========================

_COLUMN_HINTS = {
    "account_number": (
        "Account Number column not found.\n\n"
        "Please make sure your file has a column named something like:\n"
        "  • Account Number\n"
        "  • Account No\n"
        "  • Acct No\n\n"
        "Rename the column and try again."
    ),
    "bank": (
        "Bank column not found.\n\n"
        "Please make sure your file has a column named something like:\n"
        "  • Bank\n"
        "  • Bank Name\n"
        "  • Financial Institution\n\n"
        "Rename the column and try again."
    ),
    "submitted_name": (
        "Account holder name column not found.\n\n"
        "Cross-check mode compares names, so your file needs a column like:\n"
        "  • Account Name\n"
        "  • Name\n"
        "  • Employee Name\n"
        "  • Customer Name\n\n"
        "Rename the column and try again, or switch to Extract mode."
    ),
}

_ROW_STATUS_MESSAGES = {
    # What the user sees in the output Excel file
    "RESOLVED":          "Verified",
    "MATCH":             "Name Matches",
    "REVIEW":            "Needs Review",
    "MISMATCH":          "Name Mismatch",
    "ACCOUNT_NOT_FOUND": "Account Not Found",
    "UNKNOWN_BANK":      "Bank Not Recognised",
    "NO_INTERNET":       "Connection Failed",
    "TIMEOUT":           "Request Timed Out",
}


def _user_status(internal_status: str) -> str:
    """Map an internal status code to the label shown in the output file."""
    return _ROW_STATUS_MESSAGES.get(internal_status, internal_status)


def process_file(
    input_path,
    output_path,
    secret_key,
    search_map,
    mode="extract",         # "extract" | "crosscheck"
    groq_api_key=None,
    max_rows=None
):
    # --------------------------------------------------
    # Load file
    # --------------------------------------------------
    try:
        df = read_file(input_path)
    except Exception:
        raise ValueError(
            "Could not open the file. "
            "Please make sure it is a valid Excel (.xlsx) or CSV file and is not open in another program."
        )

    df.columns = [str(c).strip() for c in df.columns]

    if df.empty:
        raise ValueError(
            "The file appears to be empty. "
            "Please check that it contains data rows and try again."
        )
    
    if max_rows is not None and len(df) > max_rows:
        raise ValueError(
            f"Demo mode allows a maximum of {max_rows} rows."
        )

    # --------------------------------------------------
    # Column detection — practical errors
    # --------------------------------------------------
    account_number_col = detect_column(df.columns, COLUMN_ALIASES["account_number"])
    bank_col           = detect_column(df.columns, COLUMN_ALIASES["bank"])
    submitted_name_col = detect_column(df.columns, COLUMN_ALIASES["submitted_name"])

    if not account_number_col:
        raise ValueError(_COLUMN_HINTS["account_number"])

    if not bank_col:
        raise ValueError(_COLUMN_HINTS["bank"])

    if mode == "crosscheck" and not submitted_name_col:
        raise ValueError(_COLUMN_HINTS["submitted_name"])

    # --------------------------------------------------
    # Process rows
    # --------------------------------------------------
    results = []
    total = len(df)

    for idx, row in df.iterrows():
        account_number = (
            str(row[account_number_col])
            .strip()
            .replace(".0", "")
        )
        bank = str(row[bank_col]).strip()

        print(f"[{idx + 1}/{total}] Checking {account_number} | {bank}")

        # --- Bank resolution ---
        bank_code = get_bank_code(bank, search_map, groq_api_key=groq_api_key)

        if not bank_code:
            results.append({
                "resolved_name": None,
                "score": None,
                "status": _user_status("UNKNOWN_BANK"),
                "note": f"Could not recognise \"{bank}\" as a valid Ghanaian bank."
            })
            continue

        # --- Account resolution ---
        response = resolve_account(account_number, bank_code, secret_key)
        resolved_name = response["account_name"]

        if response["success"]:
            if mode == "crosscheck":
                submitted_name = row[submitted_name_col]
                score = fuzz.token_sort_ratio(
                    normalize_name(submitted_name),
                    normalize_name(resolved_name)
                )
                internal = classify_score(score)

                note_map = {
                    "MATCH":    "Name matches the bank record.",
                    "REVIEW":   "Name is similar but not exact — please verify manually.",
                    "MISMATCH": "Name does not match the bank record.",
                }
                note = note_map[internal]
            else:
                score    = None
                internal = "RESOLVED"
                note     = "Account verified successfully."

        else:
            score = 0
            error = response["error"]

            if error == "NO_INTERNET":
                internal = "NO_INTERNET"
                note     = "No internet connection at the time of this check. Please re-run when online."
            elif error == "TIMEOUT":
                internal = "TIMEOUT"
                note     = "The request took too long. This is usually temporary — re-running should fix it."
            else:
                internal = "ACCOUNT_NOT_FOUND"
                note     = (
                    f"Account {account_number} could not be found at {bank}. "
                    "Double-check the account number and bank name."
                )

        results.append({
            "resolved_name": resolved_name,
            "score":         score,
            "status":        _user_status(internal),
            "note":          note,
        })

        time.sleep(0.3)

    # --------------------------------------------------
    # Build output DataFrame
    # --------------------------------------------------
    results_df = pd.DataFrame(results)

    if mode == "extract":
        results_df = results_df.drop(columns=["score"])

    final_df = pd.concat([df, results_df], axis=1)
    final_df.to_excel(output_path, index=False)

    _style_workbook(output_path, mode)

    print("\nDone!")


# =========================
# EXCEL STYLING
# =========================

# Map the user-facing status labels to fill colours
_STATUS_COLORS = {
    "Verified":          ("C6EFCE", "C6EFCE"),   # green
    "Name Matches":      ("C6EFCE", "C6EFCE"),   # green
    "Needs Review":      ("FFEB9C", "FFEB9C"),   # yellow
    "Name Mismatch":     ("FFC7CE", "FFC7CE"),   # red
    "Account Not Found": ("F4B084", "F4B084"),   # orange
    "Bank Not Recognised": ("D9D9D9", "D9D9D9"), # grey
    "Connection Failed": ("D9D9D9", "D9D9D9"),   # grey
    "Request Timed Out": ("D9D9D9", "D9D9D9"),   # grey
}


def _style_workbook(output_path, mode):
    wb = load_workbook(output_path)
    ws = wb.active

    # Header
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Column widths
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(column)].width = min(max_length + 5, 60)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Status column — conditional formatting using user-facing labels
    status_col = None
    for cell in ws[1]:
        if cell.value == "status":
            status_col = cell.column
            break

    if status_col:
        col_letter  = get_column_letter(status_col)
        data_range  = f"{col_letter}2:{col_letter}{ws.max_row}"

        for label, (start_color, end_color) in _STATUS_COLORS.items():
            ws.conditional_formatting.add(
                data_range,
                CellIsRule(
                    operator="equal",
                    formula=[f'"{label}"'],
                    fill=PatternFill(
                        start_color=start_color,
                        end_color=end_color,
                        fill_type="solid"
                    )
                )
            )

    # Score column formatting (crosscheck only)
    if mode == "crosscheck":
        score_col = None
        for cell in ws[1]:
            if cell.value == "score":
                score_col = cell.column
                break
        if score_col:
            col_letter = get_column_letter(score_col)
            for row in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row}"].number_format = "0.00"

    wb.save(output_path)
